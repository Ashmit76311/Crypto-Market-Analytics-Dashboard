import os
import warnings
import mysql.connector
import pandas as pd

warnings.filterwarnings("ignore", message="pandas only supports SQLAlchemy")
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from datetime import datetime
from dotenv import load_dotenv

load_dotenv()

DB_CONFIG = {
    "host":     os.getenv("DB_HOST", "localhost"),
    "port":     int(os.getenv("DB_PORT", 3306)),
    "user":     os.getenv("DB_USER", "root"),
    "password": os.getenv("DB_PASSWORD", "pass"),
    "database": os.getenv("DB_NAME", "crypto_db"),
}

# --- Fetch latest snapshot per coin ---
def fetch_latest() -> pd.DataFrame:
    conn = mysql.connector.connect(**DB_CONFIG)
    query = """
        SELECT m1.coin_id, m1.symbol, m1.price_usd, m1.volume_24h, m1.market_cap, m1.fetched_at
        FROM market_data m1
        INNER JOIN (
            SELECT coin_id, MAX(fetched_at) AS latest
            FROM market_data
            GROUP BY coin_id
        ) m2 ON m1.coin_id = m2.coin_id AND m1.fetched_at = m2.latest
        ORDER BY m1.market_cap DESC
    """
    df = pd.read_sql(query, conn)
    conn.close()
    return df

# --- Styles ---
HEADER_FILL   = PatternFill("solid", fgColor="1E1E2E")
HEADER_FONT   = Font(bold=True, color="00D4FF", size=11)
ROW_FILL_ODD  = PatternFill("solid", fgColor="2A2A3E")
ROW_FILL_EVEN = PatternFill("solid", fgColor="1E1E2E")
WHITE_FONT    = Font(color="FFFFFF", size=10)
CENTER        = Alignment(horizontal="center", vertical="center")
LEFT          = Alignment(horizontal="left",   vertical="center")
THIN_BORDER   = Border(
    left=Side(style="thin", color="3A3A5E"),
    right=Side(style="thin", color="3A3A5E"),
    top=Side(style="thin", color="3A3A5E"),
    bottom=Side(style="thin", color="3A3A5E"),
)

def style_header(cell, text):
    cell.value     = text
    cell.fill      = HEADER_FILL
    cell.font      = HEADER_FONT
    cell.alignment = CENTER
    cell.border    = THIN_BORDER

def style_row(cell, value, row_idx, align=CENTER):
    cell.value     = value
    cell.fill      = ROW_FILL_ODD if row_idx % 2 == 0 else ROW_FILL_EVEN
    cell.font      = WHITE_FONT
    cell.alignment = align
    cell.border    = THIN_BORDER

# --- Build Report ---
def build_report(df: pd.DataFrame):
    wb = Workbook()
    ws = wb.active
    ws.title = "Crypto Snapshot"
    ws.sheet_view.showGridLines = False

    # Title row
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value     = f"Crypto Market Snapshot  —  {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    title_cell.fill      = PatternFill("solid", fgColor="0D0D1A")
    title_cell.font      = Font(bold=True, color="00D4FF", size=13)
    title_cell.alignment = CENTER
    ws.row_dimensions[1].height = 30

    # Headers
    headers = ["#", "Coin", "Symbol", "Price (USD)", "24h Volume (USD)", "Market Cap (USD)"]
    for col, h in enumerate(headers, start=1):
        style_header(ws.cell(row=2, column=col), h)
    ws.row_dimensions[2].height = 22

    # Data rows
    for i, row in enumerate(df.itertuples(index=False), start=1):
        r = i + 2
        ws.row_dimensions[r].height = 20
        style_row(ws.cell(r, 1), i, i)
        style_row(ws.cell(r, 2), row.coin_id.title(), i, LEFT)
        style_row(ws.cell(r, 3), row.symbol.upper(), i)
        style_row(ws.cell(r, 4), f"${float(row.price_usd):,.4f}", i)
        style_row(ws.cell(r, 5), f"${float(row.volume_24h):,.0f}", i)
        style_row(ws.cell(r, 6), f"${float(row.market_cap):,.0f}", i)

    # Column widths
    col_widths = [5, 18, 10, 18, 24, 24]
    for col, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Save
    os.makedirs("reports", exist_ok=True)
    filename = f"reports/crypto_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    print(f"Report saved: {filename}")
    return filename

if __name__ == "__main__":
    df = fetch_latest()
    print(f"Fetched {len(df)} coins from DB.")
    build_report(df)
